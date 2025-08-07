from flask import Flask, render_template, request
from openpyxl import load_workbook, Workbook
import os
import pandas as pd

app = Flask(__name__)
PASSWORD = "bem2025"

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    nama = request.form['nama']
    kelas = request.form['kelas']
    NIM = request.form['NIM']
    pilihan = request.form['pilihan']

    if not os.path.exists("data.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.append(["Nama", "Kelas","NIM","Pilihan"])
        wb.save("data.xlsx")

    wb = load_workbook("data.xlsx")
    ws = wb.active
    ws.append([nama, kelas, NIM, pilihan])
    wb.save("data.xlsx")

    return render_template('success.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        password = request.form['password']
        if password == PASSWORD:
            df = pd.read_excel("data.xlsx")
            return render_template('result.html', tables=[df.to_html(index=False)])
        else:
            return "Password salah!"
    return render_template('login.html')

if __name__ == '__main__':
    app.run(debug=True)
